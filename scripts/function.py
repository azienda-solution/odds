   
import csv
import io
import json
import math
import mimetypes
import os
import pickle
import re
import subprocess
from playsound import playsound
import unicodedata
from urllib.parse import urlparse
from urllib.request import urlopen
import requests
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl

from logging import exception
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import InvalidSelectorException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support.ui import Select
from webdriver_manager.firefox import GeckoDriverManager
################

import sys
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException

from webdriver_manager.chrome import ChromeDriverManager
import undetected_chromedriver as uc

import time
from datetime import datetime

from env import API_KEY, BASE_URL, COMMENCE_TIME_FROM, COMMENCE_TIME_TO, DIFFERENC, MARKETS, OPENAI_KEY, REGIONS
last_check_time = time.time()

def append_new_line(file_name, text_to_append):
    # Check if the directory exists, if not, create it
    directory = os.path.dirname(file_name)
    if directory == '':
        directory = '.'
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Append the new line to the file
    with open(file_name, "a+", encoding="utf-8", errors="ignore") as file_object:
        file_object.seek(0)
        data = file_object.read(100)
        if len(data) > 0:
            file_object.write("\n")
        file_object.write(text_to_append)
        
def can_open_yt(url):
    try:
        res = requests.get(url, timeout=10)
        resp = urlopen(url)
        return True
    except Exception as e:
        print(e)
        return False
    
    
# fonction pour donner du délai et cliquer les xpath
def waitBeforeClickOnXpath(driver, xPath):
    time.sleep(1)
    print("clicking on " + xPath + "...")
    button = driver.find_element(By.XPATH, xPath)
    driver.execute_script("arguments[0].click();", button)
    time.sleep(1)
    print("Continue the script")

def waitBeforeClickOnClass(driver, className):
    print("waiting page loading")
    time.sleep(3)
    print("clicking on " + className + "...")
    button = driver.find_element(By.CLASS_NAME, className)
    driver.execute_script("arguments[0].click();", button)
    print("button clicked")
    print("now waiting server response..")
    time.sleep(3)
    print("Continue the script")

def waitBeforeClickOnId(driver, id):
    print("waiting page loading")
    time.sleep(3)
    print("clicking on " + id + "...")
    button = driver.find_element(By.ID, id)
    driver.execute_script("arguments[0].click();", button)
    print("button clicked")
    print("now waiting server response..")
    time.sleep(3)
    print("Continue the script")

# rempli de texte la case formulaire avec l'id correspondant
def fillById(driver, id, filler):
    print("waiting page loading")
    time.sleep(3)
    driver.find_element(By.ID, id).send_keys(filler)
    print("form filled")
    print("now waiting server response..")
    time.sleep(3)
    print("Continue the script")

def fillByIdWithSteps(driver, id ,filler):
    print("waiting page loading")
    time.sleep(3)
    driver.find_element(By.ID, id).send_keys(Keys.CONTROL + "a")
    print("Taking all that already exist")
    time.sleep(1)
    driver.find_element(By.ID, id).send_keys(Keys.DELETE)
    print("Cleaning")
    time.sleep(1)
    driver.find_element(By.ID, id).send_keys(filler)
    print("Fill with our value")
    time.sleep(1)
    print("Complete")
    print("now waiting server response..")
    time.sleep(3)
    print("Continue the script")

def fillByClass(driver, clss ,filler):
    print("waiting page loading")
    time.sleep(3)
    element = driver.find_element_by_class_name(clss).click()
    time.sleep(1)
    element.send_keys(filler)
    print("Fill with our value")
    time.sleep(1)
    print("Complete")
    print("now waiting server response..")
    time.sleep(3)
    print("Continue the script")

def fillByXpath(driver, xpath, filler):
    print("waiting page loading")
    time.sleep(3)
    driver.find_element(By.XPATH, xpath).send_keys(filler)
    time.sleep(3)
    print("Continue the script")

def tryAndRetryClickXpath(driver, xPath):
    try : 
        waitBeforeClickOnXpath(driver, xPath)
    except NoSuchElementException:
        print("the element needs to be charged...")
        time.sleep(10)
        waitBeforeClickOnXpath(driver, xPath)

def tryAndRetryClickClassName(class_name):
    try : 
        waitBeforeClickOnClass(class_name)
    except NoSuchElementException:
        print("the element needs to be charged...")
        time.sleep(10)
        waitBeforeClickOnClass(class_name)

def tryAndRetryClickID(driver, id):
    try : 
        waitBeforeClickOnClass(driver, id)
    except NoSuchElementException:
        print("the element needs to be charged...")
        time.sleep(10)
        waitBeforeClickOnClass(driver, id)


def tryAndRetryFillById(driver, id, value):
    try:
        fillById(driver,id, value)
    except NoSuchElementException:
        print("the element needs to be charged...")
        time.sleep(10)
        fillById(driver,id, value)

def tryAndRetryFillByIdWithSteps(driver, idStep1, id, value):
    try:
        button = driver.find_element(By.ID, idStep1)
        driver.execute_script("arguments[0].click();", button)
        fillById(id, value)
    except NoSuchElementException:
        button = driver.find_element(By.ID, idStep1)
        driver.execute_script("arguments[0].click();", button)
        print("the element needs to be charged...")
        time.sleep(10)
        fillById(id, value)
    except ElementNotInteractableException:
        button = driver.find_element(By.ID, idStep1)
        driver.execute_script("arguments[0].click();", button)
        print("the element needs to be charged...")
        time.sleep(10)
        fillById(id, value)

def writeLetterByLetterId(driver, id, word):
    print("waiting page loading")
    time.sleep(3)
    driver.find_element(By.ID, id).send_keys(Keys.CONTROL + "a")
    print("Taking all that already exist")
    time.sleep(1)
    driver.find_element(By.ID, id).send_keys(Keys.DELETE)
    print("Cleaning")
    for i in word:
        driver.find_element(By.ID, id).send_keys(i)
        
def getinnertextXpath(driver, xPath):
    try:
        result = ""
        result = driver.find_element(By.XPATH, xPath)
        result = (result.get_attribute('innerText'))
    except NoSuchElementException:  #spelling error making this code not work as expected
        try:
            result = ""
            result = driver.find_elements(By.XPATH, xPath)
            result = ' '.join([span.get_attribute('innerText') for span in result])
        except NoSuchElementException:  
            
            result = ""
            pass
    return str(result)


    
def tryAndRetryFillByIdWithExtraSteps(driver, idStep1, id, value):
    try:
        button = driver.find_element(By.ID, idStep1)
        driver.execute_script("arguments[0].click();", button)
        writeLetterByLetterId(id, value)
    except NoSuchElementException:
        button = driver.find_element(By.ID, idStep1)
        driver.execute_script("arguments[0].click();", button)
        print("the element needs to be charged...")
        time.sleep(10)
        writeLetterByLetterId(id, value)
    except ElementNotInteractableException:
        button = driver.find_element(By.ID, idStep1)
        driver.execute_script("arguments[0].click();", button)
        print("the element needs to be charged...")
        time.sleep(10)
        writeLetterByLetterId(id, value)

def tryAndRetryFillByXpath(driver, xpath, value):
    try:
        fillByXpath(driver, xpath, value)
    except NoSuchElementException:
        print("the element needs to be charged...")
        time.sleep(5)
        tryAndRetryFillByXpath(driver, xpath, value)

def if_as_value_FillByXpath(driver, xpath, value):
    if(len(value) > 2):
        try:
            fillByXpath(driver, xpath, value)
        except NoSuchElementException:
            pass
    else:
        pass

    
    
def american_to_european_odds(american_odds):
    if american_odds < 0:
        return 1 + (100 / abs(american_odds))
    else:
        return 1 + (american_odds / 100)
    
def check_price_difference(events):
    result = []
    for event in events:
        home_team = event["home_team"]
        away_team = event["away_team"]
        for bookmaker in event["bookmakers"]:
            for market in bookmaker["markets"]:
                outcomes = market["outcomes"]
                prices = []
                for outcome in outcomes:
                    price = outcome["price"]
                    # Vérifier le format des cotes et convertir si nécessaire
                    if price > -100 and price < 100:
                        prices.append(price)  # Cotes européennes
                    else:
                        prices.append(american_to_european_odds(price))  # Convertir les cotes américaines en cotes européennes
                
                if abs(prices[0] - prices[1]) > DIFFERENC:
                    result.append({
                        "home_team": home_team,
                        "away_team": away_team,
                        "bookmaker": bookmaker["title"],
                        "market": market["key"],
                        "home_team_price": prices[1],
                        "away_team_price": prices[0]
                    })
    return result


def download_video(url, save_path):
    # Send a GET request to the video URL
    response = requests.get(url, stream=True)
    
    # Raise an exception if the request was unsuccessful
    response.raise_for_status()
    
    # Get the Content-Type from the response headers
    content_type = response.headers.get('Content-Type')
    
    # Guess the file extension from the Content-Type
    extension = mimetypes.guess_extension(content_type) if content_type else '.mp4'
    
    # Get the video file name from the URL and ensure it has the correct extension
    video_filename = url.split('/')[-1].split('?')[0]
    if not video_filename.endswith(extension):
        video_filename += extension
    
    # Create the full path to save the video
    full_path = os.path.join(save_path, video_filename)
    directory = os.path.dirname(full_path)
    if not os.path.exists(directory):
        os.makedirs(directory)
    
    # Write the content of the response to a file
    with open(full_path, 'wb') as video_file:
        for chunk in response.iter_content(chunk_size=8192):
            video_file.write(chunk)
    
    # Return the path of the saved video
    return full_path

def get_content_type(file_name):
    mime_type, _ = mimetypes.guess_type(file_name)
    if mime_type is not None:
        return mime_type
    return 'application/octet-stream'

def get_extension_from_mime_type(mime_type):
    mime_to_extension = {
        'image/jpeg': 'jpg',
        'image/png': 'png',
        'video/mp4': 'mp4',
        'video/webm': 'webm',
        'video/x-matroska': 'mkv',
        'video/x-msvideo': 'avi',
        'application/octet-stream': 'bin'
    }
    return mime_to_extension.get(mime_type, 'bin')

def convet_to_european_odds(_odd):
    if _odd > -100 and _odd < 100:
        _odd = _odd
    else:
        _odd = american_to_european_odds(_odd)
    return _odd

def calcul_diff(home_odd, away_odd):
    home_odd = convet_to_european_odds(home_odd)
    away_odd = convet_to_european_odds(away_odd)
    
    if home_odd and away_odd:
        difference = abs(home_odd - away_odd)
        return difference

def check_price_difference_onfileJson(events):
    result = []
    for sport_key, sport_events in events.items():
        for event in sport_events:
            home_team = event["home_team"]
            away_team = event["away_team"]
            for bookmaker in event["bookmakers"]:
                for market in bookmaker["markets"]:
                    outcomes = market["outcomes"]
                    prices = []
                    for outcome in outcomes:
                        price = outcome["price"]
                        # Vérifier le format des cotes et convertir si nécessaire
                        if price > -100 and price < 100:
                            prices.append(price)  # Cotes européennes
                        else:
                            prices.append(american_to_european_odds(price))  # Convertir les cotes américaines en cotes européennes
                    
                    if abs(prices[0] - prices[1]) > DIFFERENC:
                        result.append({
                            "home_team": home_team,
                            "away_team": away_team,
                            "bookmaker": bookmaker["title"],
                            "market": market["key"],
                            "home_team_price": prices[1],
                            "away_team_price": prices[0]
                        })
    return result

def initGoogle(driver):
    driver.get('https://www.google.com/')
    waitloading(4, driver=driver)
    cookieGoogle = driver.find_element(By.ID, 'L2AGLb').click()
    try:
        driver.find_element(By.CLASS_NAME, 'h-captcha')
    except NoSuchElementException:
        print("No captcha")

    if cookieGoogle:
        print("GOOGLE a changé l'id recupere le nouveau")
    else:
        print("Init Google...")
        
def waitloading(times, driver):
    times = int(times)
    time.sleep(times)
    wait = WebDriverWait(driver, times)
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

def clear_element(driver, xpath):
    elem2 = driver.find_element(By.XPATH, xpath)
    driver.execute_script('arguments[0].value = "";', elem2)
    
def check_exists_by_xpath(driver, xpath):
    try:
        time.sleep(3)
        driver.find_element(By.XPATH, xpath)
        return 0
    except NoSuchElementException:
        return 1
    
def findATTR(driver, xpath, attr):
    try:
        value_attr = driver.find_element(By.XPATH, xpath)
        value_attr = value_attr.get_attribute(attr)
    except NoSuchElementException:
        value_attr = ''
        pass
    return str(value_attr)

def get_all_a_link_in_tag(driver, xpath, attr):
    my_list = list()
    links = driver.find_elements(By.XPATH, xpath)
    for i in links:
        step1 = (i.get_attribute(attr))
        my_list.append(step1)
    return my_list

def clean_url(url):
    # Remove the protocol (http or https) and the domain
    cleaned_url = re.sub(r'^https?://[^/]+/', '', url)
    # Remove all slashes from the remaining part of the URL
    cleaned_url = cleaned_url.replace('/', '-')
    return cleaned_url

def json_to_excel(json_data, excel_file):
    # Aplatir les données JSON
    flattened_data = []
    for match in json_data:
        home_team = match["home_team"]
        away_team = match["away_team"]
        date = match["date"]
        for bookmaker in match["bookmakers"]:
            if len(bookmaker)> 0:
                bookmaker_title = bookmaker["title"]
                for outcome in bookmaker["outcomes"]:
                    flattened_data.append({
                        "home_team": home_team,
                        "away_team": away_team,
                        "date": date,
                        "bookmaker": bookmaker_title,
                        "outcome_name": outcome["name"]  if "name" in outcome else '',
                        "odds": outcome["odds"]  if "odds" in outcome else '',
                        "difference": outcome["difference"] if "difference" in outcome else ''
                    })
    
    # Convertir les données JSON aplaties en DataFrame pandas
    df = pd.DataFrame(flattened_data)
    
    # Enregistrer le DataFrame en fichier Excel
    df.to_excel(excel_file, index=False)
    
    print(f"Data successfully written to {excel_file}")
    
def json_to_csv(json_data, csv_file):
        # Aplatir les données JSON
    flattened_data = []
    for match in json_data:
        home_team = match["home_team"]
        away_team = match["away_team"]
        date = match["date"]
        for bookmaker in match["bookmakers"]:
            bookmaker_title = bookmaker["title"]
            for outcome in bookmaker["outcomes"]:
                flattened_data.append({
                    "home_team": home_team,
                    "away_team": away_team,
                    "date": date,
                    "bookmaker": bookmaker_title,
                    "outcome_name": outcome["name"]  if "name" in outcome else '',
                    "odds": outcome["odds"]  if "odds" in outcome else '',
                    "difference": outcome["difference"] if "difference" in outcome else ''
                })
    
    # Convertir les données JSON aplaties en DataFrame pandas
    df = pd.DataFrame(flattened_data)
    
    # Enregistrer le DataFrame en fichier CSV
    df.to_csv(csv_file, index=False)
    
    print(f"Data successfully written to {csv_file}")
    
    
def convert_to_number(text):
    try:
        return int(text)
    except ValueError:
        try:
            return float(text)
        except ValueError:
            return None
       
def fetch_odds(sport):
    url = f"{BASE_URL}{sport}/odds/"
    params = {
        'apiKey': API_KEY,
        'regions': REGIONS,
        'markets': MARKETS,
        'oddsFormat': 'american',
        'commenceTimeFrom': COMMENCE_TIME_FROM,
        'commenceTimeTo': COMMENCE_TIME_TO
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Failed to fetch odds for {sport}. Status code: {response.status_code}, Response: {response.text}")
        return None

# Fetching all sports
def fetch_all_sports():
    url = f"{BASE_URL}?apiKey={API_KEY}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Failed to fetch sports. Status code: {response.status_code}, Response: {response.text}")
        return None

def json_to_excel_all(json_data, excel_file):
    # A list to hold all the rows of data
    rows = []
    
    # Iterate through each sport in the JSON data
    for sport_key, games in json_data.items():
        for game in games:
            base_info = {
                "sport_key": game["sport_key"],
                "sport_title": game["sport_title"],
                "commence_time": game["commence_time"],
                "home_team": game["home_team"],
                "away_team": game["away_team"]
            }
            for bookmaker in game["bookmakers"]:
                for market in bookmaker["markets"]:
                    for outcome in market["outcomes"]:
                        row = {
                            **base_info,
                            "bookmaker_key": bookmaker["key"],
                            "bookmaker_title": bookmaker["title"],
                            "market_key": market["key"],
                            "outcome_name": outcome["name"],
                            "price": convet_to_european_odds(convert_to_number(outcome["price"]))
                        }
                        # Include point if it exists
                        if "point" in outcome:
                            row["point"] = outcome["point"]
                        else:
                            row["point"] = None
                        rows.append(row)
    
    # Convert list of rows to DataFrame
    df = pd.DataFrame(rows)
    
    # Save DataFrame to Excel file
    df.to_excel(excel_file, index=False)
    
    print(f"Data successfully written to {excel_file}")
    
def scrap_selenium():
    option = FirefoxOptions()
    option.add_argument('--disable-notifications')
    option.add_argument("--mute-audio")
    #option.add_argument("--headless")
    option.add_argument("user-agent=Mozilla/5.0 (iPhone; CPU iPhone OS 10_3 like Mac OS X) AppleWebKit/602.1.50 (KHTML, like Gecko) CriOS/56.0.2924.75 Mobile/14E5239e Safari/602.1")
    
    driverinstance = webdriver.Firefox(executable_path=GeckoDriverManager().install(), options=option)
    driverinstance.set_page_load_timeout(12)
    driverinstance.maximize_window()
    initGoogle(driverinstance)
    waitloading(2, driver=driverinstance)
    return driverinstance


def check_and_load_cookies(driver, cookie_file="cookies.pkl", sound_file="D:/Documents/Advanced-Python/ODDS/config/fail.mp3"):
    # Check if the cookie file exists
    if os.path.exists(cookie_file):
        print("Cookie file found. Loading cookies...")
        # Load cookies from the file and add them to the browser
        cookies = pickle.load(open(cookie_file, "rb"))
        for cookie in cookies:
            driver.add_cookie(cookie)
        print("Cookies loaded successfully.")
    else:
        print("Cookie file not found. Playing alert sound and saving current cookies...")
        # Play a sound to alert that cookies do not exist
        playsound(sound_file)
        
        # Save the cookies from the current session
        pickle.dump(driver.get_cookies(), open(cookie_file, "wb"))
        print("Cookies saved successfully.")

def scrap_selenium_v1(init_url):
    """
    
    VERSION 1
    ______________________________________________2022-2024
    
    option = FirefoxOptions()
    option.add_argument('--disable-notifications')
    option.add_argument("--mute-audio")
    #option.add_argument("--headless")
    option.add_argument("user-agent=Mozilla/5.0 (iPhone; CPU iPhone OS 10_3 like Mac OS X) AppleWebKit/602.1.50 (KHTML, like Gecko) CriOS/56.0.2924.75 Mobile/14E5239e Safari/602.1")

    driverinstance = webdriver.Firefox(options=option, executable_path='D:/Documents/bin/geckodriver/geckodriver')
    initGoogle(driverinstance)
    waitloading(2, driver=driverinstance)
    return driverinstance"""
    """

    chrome_options = Options()
    
    # Add necessary Chrome options
    # chrome_options.add_argument('--headless')  # Uncomment if you want to run Chrome in headless mode
    chrome_options.add_argument('disable-infobars')
    chrome_options.add_argument('--disable-gpu')  # Disable GPU acceleration if using headless
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("user-data-dir=selenium")
    
    # Use Service to specify the ChromeDriver path
    chrome_service = Service(executable_path='D:/Documents/bin/geckodriver/chromedriver')

    # Initialize the WebDriver
    driverinstance = webdriver.Chrome(service=chrome_service, options=chrome_options)"""
    
    """
    VERSION 2
    _____________________________________________________2024-2024
    
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    #chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--profile-directory=Default')
    #chrome_options.add_argument('--user-data-dir=~/.config/google-chrome')
    #chrome_options.add_argument("user-data-dir=selenium")
    driverinstance = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

    initGoogle(driverinstance)
    waitloading(2, driver=driverinstance)
    #check_and_load_cookies(driverinstance)
    return driverinstance"""
    
    
    """
    VERSION 3
    _____________________________________________________2025-2025
    """
    # pip3 install seleniumbase
    from seleniumbase import Driver

    # initialize the driver in GUI mode with UC enabled
    driver = Driver(uc=True, headless=True)

    initGoogle(driver)
    waitloading(2, driver=driver)
    driver.uc_open_with_reconnect(init_url, reconnect_time=6)
    driver.uc_gui_click_captcha()
    return driver

    """driver.save_screenshot("cloudflare-challenge.png")
    driver.quit()"""
    
    
def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

# Function to parse the HTML file and extract the data
def parse_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')
        
    games = []
    prints = []
    current_date = None
    match_info = {}
    today_date = datetime.now().strftime('%Y-%m-%d')

    for row in soup.select('#odds_tb tr'):
        cells = row.find_all('td')
        
        # Check for date row
        if len(cells) == 1 and 'colspan' in cells[0].attrs:
            current_date = cells[0].get_text(strip=True)
            print(f"Found date: {current_date}")
            continue
        
        # Process match rows
        if len(cells) == 8:
            initial_odds_home = cells[3].get_text(strip=True)
            initial_draw_odds = cells[4].get_text(strip=True)
            initial_odds_away = cells[5].get_text(strip=True)
            
            if is_float(initial_odds_home) and is_float(initial_draw_odds) and is_float(initial_odds_away):
                match_info = {
                    'home_team': cells[2].get_text(strip=True),
                    'away_team': cells[6].get_text(strip=True),
                    'date': current_date,
                    'initial_odds_home': float(initial_odds_home),
                    'initial_draw_odds': float(initial_draw_odds),
                    'initial_odds_away': float(initial_odds_away),
                    'initial_difference': abs(float(initial_odds_home) - float(initial_odds_away))
                }
                #print(f"Initial match info: {match_info}")
        
        # Process odds update rows
        elif len(cells) == 3 and match_info:
            actual_odds_home = cells[0].get_text(strip=True)
            draw_odds_actual = cells[1].get_text(strip=True)
            actual_odds_away = cells[2].get_text(strip=True)
            
            if is_float(actual_odds_home) and is_float(draw_odds_actual) and is_float(actual_odds_away):
                match_info.update({
                    'actual_odds_home': float(actual_odds_home),
                    'draw_odds_actual': float(draw_odds_actual),
                    'actual_odds_away': float(actual_odds_away),
                    'actual_difference': abs(float(actual_odds_home) - float(actual_odds_away))
                })
                if match_info["initial_difference"] > DIFFERENC or match_info["actual_difference"] > DIFFERENC:
                    prints.append(match_info.copy())
                games.append(match_info.copy())  # Use copy to ensure we don't overwrite previous entries
                #print(f"Updated match info: {match_info}")
                match_info = {}  # Reset match_info for the next match
    
    print(f"Total games found: {len(games)}")
    if len(prints) > 0:
        df = pd.DataFrame(prints)
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', None)
        # Affichage du DataFrame
        print(df)
        file_path = os.path.join('daily_data', str(today_date), f'{today_date}.txt')
        append_new_line(file_path, str(df))
    return games

def array_to_csv_string(array):
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(array)
    return output.getvalue().strip()

def oddspedia(file_path):
        
    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    # Trouver tous les blocs de match
    match_blocks = soup.find_all('div', class_='match-list-item')

    # Liste pour stocker les informations de chaque match
    matches_info = []

    for match in match_blocks:
        today_date = datetime.now().strftime('%Y-%m-%d')
        first_link = match.find('a', href=True)
        if first_link:
            href_value = first_link['href']
            category = href_value.split('/')[1] if len(href_value) > 1 else ''
        else:
            category = ''
    
        # Extraire les noms des équipes
        team_names = match.find_all('div', class_='match-team__name')
        if len(team_names) >= 2:
            home_team = team_names[0].get_text(strip=True)
            away_team = team_names[1].get_text(strip=True)
        
            # Extraire les cotes
            odds_values = match.find_all('span', class_='odd__value')
            if len(odds_values) < 2:
                continue
            
            if len(odds_values) == 3:
                odds_home = str(odds_values[0].get_text(strip=True))
                odds_draw = str(odds_values[1].get_text(strip=True))
                odds_away = str(odds_values[2].get_text(strip=True))
            
                # Créer un dictionnaire avec les informations du match
                match_info = {
                    'home_team': home_team,
                    'away_team': away_team,
                    'date': today_date,
                    'initial_odds_home': odds_home,
                    'initial_draw_odds': odds_draw,
                    'initial_odds_away': odds_away,
                    'initial_difference': abs(float(odds_home) - float(odds_away)),
                    'actual_odds_home': "",
                    'draw_odds_actual': "",
                    'actual_odds_away': "",
                    'actual_difference': "",
                    'category': category
                }
                matches_info.append(match_info)
            if len(odds_values) == 2:
                odds_home = str(odds_values[0].get_text(strip=True))
                odds_away = str(odds_values[1].get_text(strip=True))
                odds_draw = ''
                # Créer un dictionnaire avec les informations du match
                match_info = {
                    'home_team': home_team,
                    'away_team': away_team,
                    'date': today_date,
                    'initial_odds_home': odds_home,
                    'initial_draw_odds': odds_draw,
                    'initial_odds_away': odds_away,
                    'initial_difference': abs(float(odds_home) - float(odds_away)),
                    'actual_odds_home': "",
                    'draw_odds_actual': "",
                    'actual_odds_away': "",
                    'actual_difference': "",
                    'category': category
                }
                matches_info.append(match_info)
    return matches_info
    
def merge_result(parent, child):
    for elm in child:
        parent.append(elm)
    return parent

def check_number_type(value):
    if isinstance(value, float):
        return "float"
    elif isinstance(value, int):
        return "int"
    else:
        return "string" 

def oddsportal(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    cleaned_lines = []
    skip_next = False

    for i in range(len(lines)):
        line = lines[i].strip()

        # Si on doit sauter la ligne, on continue
        if skip_next:
            skip_next = False
            continue

        if (not is_float(line) and len(line) > 2 and 
            i + 1 < len(lines) and line == lines[i + 1].strip()):
            skip_next = True  # On saute la prochaine ligne (on garde la première)
        
        cleaned_lines.append(line)

    cleaned_text = "\n".join(cleaned_lines)
    cleaned_text = cleaned_text.replace("\n\n", "\n")
    cleaned_text = "\n".join([line for line in cleaned_text.split("\n") if line.strip() != ""])

    match_info_list = []

    lines = cleaned_text.splitlines()

    for i in range(len(lines)):
        line = lines[i].strip()
        if "–" in line:
            parts = line.split("–")
            if len(parts) == 2:
                home_team = lines[i-1].strip()
                away_team = lines[i+1].strip()

                if len(home_team) > 3 and len(away_team) < 3:
                    home_team = lines[i-1].strip()
                    away_team = lines[i+2].strip()
                # Vérifie qu'il y a au moins trois lignes suivantes pour les cotes
                if i + 3 < len(lines):
                    odds_home = lines[i + 2].strip()
                    odds_draw = lines[i + 3].strip()
                    odds_away = lines[i + 4].strip()
                    
                    if 'int' in check_number_type(odds_away) and 'float' in check_number_type(odds_home) and 'float' in check_number_type(odds_draw):
                    
                        odds_home = odds_home
                        odds_away = odds_draw
                        odds_draw = 0

                    try:
                        odds_home = float(odds_home)
                    except ValueError:
                        odds_home = lines[i + 3].strip()
                        odds_away = lines[i + 4].strip()
                        odds_draw = 0
                    try:
                        odds_home = float(odds_home)
                        odds_draw = float(odds_draw)
                        odds_away = float(odds_away)
                        initial_difference = abs(float(odds_home) - float(odds_away))
                    except ValueError:
                        odds_home = ""
                        odds_draw = ""
                        odds_away = ""
                        initial_difference = ""

                    match_info = {
                        'home_team': home_team,
                        'away_team': away_team,
                        'date': "-",
                        'initial_odds_home': odds_home,
                        'initial_draw_odds': odds_draw,
                        'initial_odds_away': odds_away,
                        'initial_difference': initial_difference,
                        'actual_odds_home': "",
                        'draw_odds_actual': "",
                        'actual_odds_away': "",
                        'actual_difference': ""
                    }

                    match_info_list.append(match_info)

    return match_info_list

            
def save_to_excel(games, excel_file):
    df = pd.DataFrame(games)
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    counter = 0
    while True:
        try:
            # Adjust sheet name if there's a conflict
            sheet_name = today_date if counter == 0 else f"{today_date}-{counter}"
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Data saved to sheet: {sheet_name}")
            break
        except Exception as e:
            print(f"Error: {e}. Retrying with a new sheet name.")
            counter += 1

def check_and_refresh(driver, expected_url, timeout=120):
    """
    Check if the time passed since the last URL load exceeds `timeout` seconds
    or if the current URL differs from `expected_url`. Refresh the page if needed.
    
    :param driver: Selenium WebDriver instance
    :param expected_url: The expected URL to check against `driver.current_url`
    :param timeout: Timeout in seconds (default: 60)
    """
    global last_check_time
    current_time = time.time()
    
    if (current_time - last_check_time > timeout) and (driver.current_url == expected_url):
        print("Condition met: refreshing the page...")
        driver.refresh()
        last_check_time = time.time()
        
def check_fitness_state(result):
    # Define the possible states
    possible_states = ['fit', 'neutral', 'injury']
    for state in possible_states:
        if state in result:
            return state
    return ""

def remove_accents(text):
    # Normalize the text to decompose accents and special characters
    text = unicodedata.normalize('NFKD', text)
    
    # Remove the accents (diacritical marks)
    text = text.encode('ASCII', 'ignore').decode('utf-8')
    
    # Optionally, remove any other non-alphanumeric characters (if needed)
    text = re.sub(r'[^\w\s]', '', text)
    
    return text

def extract_content(driverinstance, google_query):
    my_list = fing_google(driverinstance, google_query)
    return my_list

def flashscore_player(driverinstance, team, country, sport):
    list_player = []
    list_link = extract_list_from_google(driverinstance, "Squad Player of "+team + " " + sport +" " + country +" in flashscore.com")
    if len(list_link) > 1:
        for l in list_link:
            if 'flashscore' in l and 'squad' in l:
                driverinstance.get(l)
                waitloading(4, driver=driverinstance)
                try:
                    player_name_elements = driverinstance.find_elements(By.XPATH, "//div[@class='lineupTable lineupTable--soccer']//a[contains(@class, 'lineupTable__cell--name')]")
                    # Iterate through all found elements and extract the text (player name)
                    for player in player_name_elements:
                        player_name = player.text.strip()  # .text will extract the visible text
                        if len(player_name) > 1:  # Only add if there is meaningful content
                            list_player.append(player_name)
                except Exception as e:
                    print(f"An error occurred: {e}")
                return list_player
            else:
                if 'flashscore.com' in l:
                    driverinstance.get(l)
                    waitloading(4, driver=driverinstance)
                    if driverinstance.find_elements(By.XPATH, "//body"):
                        l = l + '/squad/'
                        l = l.replace('//', '/')
                        driverinstance.get(l)
                        waitloading(4, driver=driverinstance)
                        if driverinstance.find_elements(By.XPATH, "//body"):
                            try:
                                player_name_elements = driverinstance.find_elements(By.XPATH, "//div[@class='lineupTable lineupTable--soccer']//a[contains(@class, 'lineupTable__cell--name')]")
                                # Iterate through all found elements and extract the text (player name)
                                for player in player_name_elements:
                                    player_name = player.text.strip()  # .text will extract the visible text
                                    if len(player_name) > 1:  # Only add if there is meaningful content
                                        list_player.append(player_name)
                            except Exception as e:
                                print(f"An error occurred: {e}")
                            return list_player
    return list_player
    
def add_row_sheet(sheet, array_row):
    # Open an existing workbook
    try:
        wb = openpyxl.load_workbook(sheet)
    except Exception as e:
        print(f"Failed to open Excel file: {e}")
        return

    # Select the active sheet (or specify the sheet name)
    ws = wb.active

    # Add the new row data to the next empty row
    next_row = ws.max_row + 1
    for col_num, value in enumerate(array_row, start=1):  # Columns start from 1 (A)
        ws.cell(row=next_row, column=col_num, value=value)

    # Save the updated workbook
    try:
        wb.save(sheet)
        print("Row added successfully!")
    except Exception as e:
        print(f"Failed to save Excel file: {e}")

def extract_list_from_google(driverinstance, title):
    driverinstance.get('https://www.google.com/')
    waitloading(4, driver=driverinstance)
    my_list = list()
    try:
        try:
            driverinstance.find_element(By.XPATH, '//textarea[contains(@maxlength, "2048")]').send_keys(title)
        except NoSuchElementException:
            driverinstance.find_element(By.XPATH, '//input[contains(@maxlength, "2048")]').send_keys(title)
        actions = ActionChains(driverinstance)
        actions.send_keys(Keys.ENTER)
        actions.perform()
        time.sleep(5)
        waitloading(4, driver=driverinstance)
        links = driverinstance.find_elements(By.XPATH,"//div[contains(@data-snhf, '0')]//a")
        to_delete = []
        for i in links:
            step1 = (i.get_attribute('href'))
            __parsed_uri = urlparse(step1)
            __result = '{uri.scheme}://{uri.netloc}/'.format(uri=__parsed_uri)
            __hostname = __result
            to_delete.append(step1)
            EXCLUDED_KEYWORDS = ["translate.g", "translate.google", ".pdf", ".docx", ".doc", ".jpeg", ".webp", ".rar", "dictionnaire", "dictionnaire.", "facebook.", "tikok.", "/download/", "/apt/", "apt/", ".txt", "amazon.", "adobe.", ".lingue", "-francais", "/traduction/", "/dictionary", "/traduction/", ".ebay", ".fnac", "/traduction/", "definitions", "/definition", "dictionary.", "twitch.", "encyclo", ".youtube", "?pdf", "pdf=", ".html", ".php", ".pptx", "?path", "path=", "/path", ".pps", "google.", ".google.", ".tiktok", "search.do?recordID", ".pensoft.net", ".virginialiving.", "/PDF/", "/object/", "PDF/", ".mnhn.fr", "gtgrecords.net", "komitid.fr", "/translate/", "linkedin.", ".parismuseescollections."]
            if any(keyword in str(step1) for keyword in EXCLUDED_KEYWORDS):
                pass
            else:
                my_list.append(step1)
    except Exception as e:
        print(f"-------- Failed get google ( crashed ) crash cause: {e}")
        raise Exception(e)
    return my_list
        
def fing_google(driverinstance, title):
    driverinstance.get('https://www.google.com/')
    waitloading(2, driver=driverinstance)
    check_and_load_cookies(driverinstance)
    my_list = list()
    content = ''
    try:
        try:
            driverinstance.find_element(By.XPATH, '//textarea[contains(@maxlength, "2048")]').send_keys(title)
        except NoSuchElementException:
            driverinstance.find_element(By.XPATH, '//input[contains(@maxlength, "2048")]').send_keys(title)
        #tryAndRetryFillByXpath(driverinstance, '//input[contains(@maxlength, "")]', title)
        
        actions = ActionChains(driverinstance)
        actions.send_keys(Keys.ENTER)
        actions.perform()
        #time.sleep(5)
        waitloading(3, driver=driverinstance)
        content = getinnertextXpath(driverinstance, '//body//div[@id="appbar"]')
        if len(content) < 300:
            content += getinnertextXpath(driverinstance, '//body//div[@id="search"]')
    except Exception as e:
        print(f"-------- Failed get google ( crashed ) crash cause: {e}")
        raise Exception(e)
    return content


def forebet(html, types=None, folder=None):
    
    content_folder = ""
    soup = BeautifulSoup(html, 'html.parser')
    if types:
        if types == "folder":
            for filename in os.listdir(folder):
                if filename.endswith('.html'):
                    file_path = os.path.join(folder, filename)
                    with open(file_path, 'r', encoding='utf-8') as file:
                        content_folder += str(file.read())
            if len(content_folder) > 22:
                soup = BeautifulSoup(content_folder, 'html.parser')
        else:
            with open(html, 'r', encoding='utf-8') as file:
                soup = BeautifulSoup(file, 'html.parser')

    matches = []
        
    for match in soup.select('div.rcnt'):
        try:
            sport = match.select_one('strong.sportfill').text.strip() if match.select_one('strong.sportfill') else ''
            
            # Team Names
            home_team = match.select_one('.homeTeam span').text.strip() if match.select_one('.homeTeam span') else ''
            away_team = match.select_one('.awayTeam span').text.strip() if match.select_one('.awayTeam span') else ''
            
            # Date
            date = match.select_one('.date_bah').text.strip() if match.select_one('.date_bah') else ''
            
            # Probabilities
            probabilities = match.select('.fprc span')
            num_probabilities = len(probabilities)
            # Set probabilities based on the number of spans
            if num_probabilities == 3:
                home_probability = probabilities[0].text.strip() if len(probabilities) > 0 else ''
                draw_probability = probabilities[1].text.strip() if len(probabilities) > 1 else ''
                away_probability = probabilities[2].text.strip() if len(probabilities) > 2 else ''
            elif num_probabilities == 2:
                home_probability = probabilities[0].text.strip() if len(probabilities) > 0 else ''
                draw_probability = ''  # No draw for basketball with only two values
                away_probability = probabilities[1].text.strip() if len(probabilities) > 1 else ''
            else:
                home_probability = draw_probability = away_probability = ''  # Default if no valid probabilities found
            
            
            # Predictions
            prediction = match.select_one('.predict_y .forepr span, .predict .forepr span')
            prediction_value = prediction.text.strip() if prediction else ''
            
            # Scores
            correct_score = match.select_one('.ex_sc').text.strip().replace('\n', ' ') if match.select_one('.ex_sc') else ''
            correct_score = format_text_score(correct_score)
            average_score = match.select_one('.avg_sc').text.strip() if match.select_one('.avg_sc') else ''
            get_odds = match.select_one('.lscrsp').text.strip() if match.select_one('.lscrsp') else ''
            
            link = ""
            link_tag = match.select_one('a.tnmscn, a[itemprop="url"]')
            if link_tag and link_tag.get('href'):
                link = str("https://www.forebet.com/"+(link_tag['href'])).replace("//", "/")
            # Compile match details
            match_info = {
                'home_team': home_team,
                'away_team': away_team,
                'date': date,
                'home_probability': to_percentage(home_probability),
                'draw_probability': to_percentage(draw_probability),
                'away_probability': to_percentage(away_probability),
                'prediction': prediction_value,
                'correct_score': correct_score,
                'average_score': average_score,
                'sport': sport,
                'initial_difference': abs(float(home_probability) - float(away_probability)),
                'get_odds': clean_text(get_odds),
                'link': link
            }
            
            matches.append(match_info)
        
        except Exception as e:
            print(f"Error parsing match: {e}")

    return matches


def click_consent(driver, language):
    time.sleep(1)
        
    try:
        btns_list = driver.find_elements(By.TAG_NAME, "button")
    except Exception as e:
        print("Error finding buttons:", e)
        btns_list = []
    for b in btns_list:
        try:
            # print(b.get_attribute('innerText'))
            has_clicked = click_consent_list(btns_list, driver, language)
            
            if not has_clicked:
                btns_list = driver.find_element(By.CLASS_NAME, "button")
                has_clicked = click_consent_list(btns_list, driver, language)
            else:
                return has_clicked
        except Exception as e:
            print(f"Error with button: {e}, skipping...")
            continue
    
    return has_clicked


TAGS_CONSENT = {'fr': {'exact': ['ok'],
                       'rel': ['accept', 'accord', 'autor', 'Autor']},
                'en': {'exact': ['ok'],
                       'rel': ['accept', 'consent', 'agree', 'autor', 'Autor']},
                }


def click_consent_list(btns_list, driver, language):
    if not btns_list:
        return False
    for btn in btns_list:
        try:
            # Ensure the button is still attached to the DOM and interactable
            assert btn.is_displayed()
            assert bool(btn.text)
            
            btn_text = btn.text.lower()
            consent_tags = TAGS_CONSENT[language]
            if any([tag in btn_text for tag in consent_tags['rel']] + 
                   [tag == btn_text for tag in consent_tags['exact']]):
                
                driver.execute_script("arguments[0].scrollIntoView();", btn)
                btn.click()
                return True
        except StaleElementReferenceException:
            continue
        except AssertionError:
            # Skip invisible or empty-text buttons
            continue
        except NoSuchElementException:
            # Handle cases where the element no longer exists
            continue
        except Exception as e:
            # Handle other unexpected errors
            continue
    return False


def is_evening():
    array_today = [
        'https://www.forebet.com/en/football-tips-and-predictions-for-today', 
                'https://www.forebet.com/en/basketball/predictions-today',
                'https://www.forebet.com/en/hockey/predictions-today',
                'https://www.forebet.com/en/american-football/predictions-today',
                'https://www.forebet.com/en/volleyball/predictions-today',
                'https://www.forebet.com/en/handball/predictions-today',
                'https://www.forebet.com/en/rugby/predictions-today']
    array_tomorrow = [
        "https://www.forebet.com/en/football-tips-and-predictions-for-tomorrow",
        "https://www.forebet.com/en/basketball/predictions-tomorrow",
        "https://www.forebet.com/en/hockey/predictions-tomorrow",
        "https://www.forebet.com/en/american-football/predictions-tomorrow",
        "https://www.forebet.com/en/volleyball/predictions-tomorrow",
        "https://www.forebet.com/en/handball/predictions-tomorrow",
        "https://www.forebet.com/en/rugby/predictions-tomorrow",
    ]
    hour = datetime.now().hour
    if hour >= 17 and hour <= 23:
        return array_tomorrow
    if hour >= 0 and hour < 15:
        return array_today

def forebet_scrap(driver):
    allcontent = str("")
    waitloading(2, driver=driver)
    click_consent(driver, 'en')
    for ii in is_evening():
        driver.get(ii)
        sport = str(ii.split("en/")[1].split("/")[0])
        sport = sport.split("-")[0] if '-' in sport else sport
        if "football" in ii:
            time.sleep(6)
        waitloading(2, driver=driver)
        if check_exists_by_xpath(driver, '//div[contains(@class, "fc-dialog-container")]//div[contains(@class, "fc-close fc-icon-button")]//span') == 0:
            tryAndRetryClickXpath(driver, '//div[contains(@class, "fc-dialog-container")]//div[contains(@class, "fc-close fc-icon-button")]//span')
        else:
            waitloading(1, driver=driver)
        content = driver.find_element(By.XPATH, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]')
        content = (content.get_attribute('outerHTML'))
        if content:
            content = ajouter_sportfill(content, sport)
        allcontent += content
    return allcontent

def ajouter_sportfill(texte, sport):
    try:
        pattern = r'(<div class="rcnt[^>]*">)(?!.*<strong class="sportfill">)'
        remplacement = r'\1 <strong class="sportfill"> '+sport+'</strong>'
        texte_modifie = re.sub(pattern, remplacement, texte)
        
        return texte_modifie

    except re.error as e:
        return texte

    except Exception as e:
        return texte

def clean_html_and_return_innertext(elements):
    try:
        # If the input is a list, process each element
        if isinstance(elements, list):
            plain_texts = " "
            for el in elements:
                html_text = el.get_attribute('innerHTML')
                soup = BeautifulSoup(html_text, 'html.parser')
                for tag in soup(['script', 'style']):
                    tag.decompose()  # Remove script and style tags
                for div in soup.find_all('div', class_='st_arrdt'):
                    div.decompose()
                plain_text = soup.get_text(strip=False)  # Get text
                plain_texts += (plain_text)
                plain_texts = clean_text(plain_texts)
            return plain_texts
        
        # If it's a single element, process it
        else:
            html_text = elements.get_attribute('innerHTML')
            soup = BeautifulSoup(html_text, 'html.parser')
            for tag in soup(['script', 'style']):
                tag.decompose()
            for div in soup.find_all('div', class_='st_arrdt'):
                div.decompose()
            plain_text = soup.get_text(strip=False)
            plain_text = clean_text(plain_text)
            return plain_text

    except Exception as e:
        return ""

def convert_sheet_csv(sheet_name, file_path):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    json_data = df.to_json(orient='records', date_format='iso')
    output_file = 'output.json'
    data = json.loads(json_data)

    return data

def clean_text(article_text):
    if not bool(article_text):
        return article_text
    article_text = re.sub(r' {2,}', ' ', article_text.replace('\xa0', ' ').replace('\n', ' ')).strip()
    cleantext = re.sub(r'\.([^. 0-9]{2})', r'. \1', article_text)
    cleanr = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')
    cleantext = re.sub(cleanr, '', cleantext)
    normalized_text = unicodedata.normalize("NFKD", cleantext)
    return normalized_text

def forebet_scrap_trend(driver, link):
    allcontent = str()
    driver.get(link)
    sport = str(link.split("en/")[1].split("/")[0])
    sport = sport.split("-")[0] if '-' in sport else sport
    waitloading(2, driver=driver)
    if check_exists_by_xpath(driver, '//div[contains(@class, "fc-dialog-container")]//div[contains(@class, "fc-close fc-icon-button")]//span') == 0:
        tryAndRetryClickXpath(driver, '//div[contains(@class, "fc-dialog-container")]//div[contains(@class, "fc-close fc-icon-button")]//span')
    else:
        waitloading(1, driver=driver)
    
    waitloading(2, driver=driver)
    content = driver.find_element(By.XPATH, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]')
    content = (content.get_attribute('outerHTML'))
    allcontent += content
            
    visited_urls = []
    pagination_container = driver.find_element(By.XPATH, '//div[contains(@class, "list-footer")]')
    pagination_links = pagination_container.find_elements(By.XPATH, './/a[@class="pagenav"]')
    
    for link in pagination_links:
        # Get the URL from the href attribute
        href = link.get_attribute('href')
        if href and href not in visited_urls:  # Avoid duplicates
            visited_urls.append(href)
            print(f"Navigating to: {href}")
            driver.get(href)
            waitloading(2, driver=driver)
            content = driver.find_element(By.XPATH, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]')
            content = (content.get_attribute('outerHTML'))
            allcontent += content
    return allcontent

def get_gpt_response_name(title, GPT_CONFIG):
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": OPENAI_KEY
    }
    data = {
        "model": "gpt-3.5-turbo-1106",
        "messages": [
            {"role": "system", "content": "You are a helpful assistant designed to output JSON."},
            {"role": "user", "content": GPT_CONFIG + title}
        ],
        "temperature": 0.7
    }

    try:
        # Send request to OpenAI API
        response = requests.post(url, headers=headers, json=data, timeout=30)
        response.raise_for_status()  # Check for HTTP errors

        # Parse the response to JSON
        response_json = response.json()

        # Extract the content from the response
        content = response_json['choices'][0]['message']['content']
        append_new_line('log_open_ai.txt', str(GPT_CONFIG)+'\n'+str(content))
        # Remove markdown formatting (triple backticks and "json" keyword)
        if content.startswith("```json") and content.endswith("```"):
            content = content.strip("```json").strip("```")

        # Unescape characters like \n, \", etc.
        unescaped_content = content.encode('utf-8').decode('unicode_escape')
        content_json = json.loads(unescaped_content)

        return content_json  # Return the array of player names

    except requests.exceptions.RequestException as e:
        print(f"HTTP Request failed: {e}")
        return None
    except (KeyError, json.JSONDecodeError) as e:
        print(f"Error parsing response: {e}")
        return None

def set_text(match__):
    text = f"""
    _______________________________________________________________________________________________________
                    - Home Team: {match__['home_team']}
                    - Away Team: {match__['away_team']}
                    - Date: {match__['date']}
                    - Home Win Probability: {match__['home_probability']}
                    - Draw Probability: {match__['draw_probability']}
                    - Away Win Probability: {match__['away_probability']}
                    - My Prediction: {match__['prediction']}
                    - Correct Score: {match__['correct_score']}
                    - Average Score: {match__['average_score']}
                    - Sport: {match__['sport']}
                    
    """
    return text

def format_text_score(text):
    if not text.isdigit():
        return text
    
    # Calculer la longueur du texte
    length = len(text)
    
    # Si la longueur est paire
    if length % 2 == 0:
        mid = length // 2
        left_part = text[:mid]
        right_part = text[mid:]
        
        # Comparer la partie gauche et droite
        if int(left_part) > int(right_part):
            return left_part + '-' + right_part
        else:
            return right_part[:len(right_part)//2] + '-' + right_part[len(right_part)//2:]
    
    # Si la longueur est impair
    else:
        mid = length // 2
        left_part = text[:mid+1]  # La partie gauche aura un caractère de plus
        right_part = text[mid+1:]
        
        # Comparer la partie gauche et droite
        if int(left_part) > int(right_part):
            return left_part + '-' + right_part
        else:
            return right_part[:len(right_part)//2] + '-' + right_part[len(right_part)//2:]

def get_trend_forebet(driver):
    html_text = ""
    try:
        divs = driver.find_elements(By.XPATH, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]//div[contains(@class, "short_trends")]')
        if divs:
            if len(divs) > 1:
                for el in divs:
                    html_text += el.get_attribute('innerText')
            else:
                html_text += divs.get_attribute('innerText')
        return html_text
    except Exception as e:
       return html_text

def cleaner(file):
    try:
        bash_command = "scripts/cleaner.sh "+file
        process = subprocess.run(bash_command, shell=True, text=True, capture_output=True)
    except Exception as e:
        pass

def to_percentage(value):
    try:
        # Convert the value to float
        float_value = float(value)
        
        # If the value is less than 1, convert to percentage
        if float_value < 1:
            return float_value * 100  # Return the value as a percentage (float)
        elif 1 <= float_value <= 100:
            return float_value  # Return as is if the value is between 1 and 100
        else:
            return float_value  # Return the value as is if it's greater than 100

    except (ValueError, TypeError):
        return value
    
def bayesian_inference(p1, p2, type_call=None):
    if str(p1) == "" or str(p2) == "":
        return ""

    if type_call is None:
        p1 = p1 / 100.0
        p2 = p2 / 100.0
    numerator = p1 * p2
    denominator = (p1 * p2) + ((1 - p1) * (1 - p2))
    result = numerator / denominator

    if type_call:
        return result
    result = to_percentage(result)
    return result


def weighted_average(p1, p2, w1=70, w2=30, type_call=None):
    if str(p1) == "" or  str(p2) == "" :
        return ""
    """
    Combines two probabilities using a weighted average.
    Weights w1 and w2 represent the confidence or reliability of each probability.
    """
    result = (w1 * p1 + w2 * p2) / (w1 + w2)
    if type_call:
        return result
    result = to_percentage(result)
    return result

def logarithmic_opinion_pool(p1, p2, type_call=None):
    if str(p1) == "" or  str(p2) == "" :
        return ""
    """
    Combines two probabilities using the logarithmic opinion pool.
    Assumes the probabilities are independent views of the event.
    """
    log_avg = (math.log(p1) + math.log(p2)) / 2
    result = math.exp(log_avg)
    if type_call:
        return result
    result = to_percentage(result)
    return result

def calcul_ponderation(source_1, source_2_api, source_3_math, source_4_math, type_call=None):
    try:
        
        if str(source_1) == "" or str(source_2_api) == "" or str(source_3_math) == "" or str(source_4_math) == "":
            return ""
        # Poids des sources : 
        poids_source_1 = 0.5  # 50% pour la source la plus fiable
        poids_source_2_api = 0.1667  # 16.67% pour la source moyennement fiable
        poids_source_3_math = 0.1667  # 16.67% pour la source moyennement fiable
        poids_source_4_math = 0.1667  # 16.67% pour la source moyennement fiable

        # Calcul de la pondération finale
        ponderation_finale = (source_1 * poids_source_1 +
                            source_2_api * poids_source_2_api +
                            source_3_math * poids_source_3_math +
                            source_4_math * poids_source_4_math)

        result = ponderation_finale
        if type_call:
            return result
        result = to_percentage(result)
        return result
    except Exception as e:
        return ""
    
def prompt(match__, last_match, trend):
    sport = match__['sport']

    def switch(sport):
        if "football" in sport:
            return """
            ### Football Specificity:
            - Focus on head-to-head history between the two teams.
            - Analyze the recent performances of the home and away teams separately.
            - Use trends such as average goals scored and conceded to evaluate consistency.
            - Consider goal differentials and the frequency of draws.
            """
        elif "american" in sport:
            return """
            ### American Football Specificity:
            - The data provided for this sport is **very robust and reliable**.
            - Trust the initial prediction provided and focus on complementing it with a small additional analysis.
            - Prioritize recent team performance and trends, such as scoring averages per game and team-specific strengths (e.g., offense vs defense efficiency).
            - Avoid unnecessary adjustments unless the trends clearly indicate an anomaly.
            """
        elif "basketball" in sport:
            return """
            ### Basketball Specificity:
            - Emphasize recent match history of each team and scoring trends.
            - Use average points per quarter and overall scoring trends to refine the forecast.
            - if in my prediction, the gap is clear between the 2 teams but in recent matches we have the impression that one team has a lot of dissuasive results when there is such a big gap, it is perhaps because their level is really different, check if they faced the same team and see the difference from this benchmark
            - Consider the team's momentum, win streaks, and defensive performance.
            """
        elif sport in ["rugby", "volleyball", "handball"]:
            return f"""
            ### {sport.capitalize()} Specificity:
            - Focus on team-level trends such as recent win/loss streaks and scoring averages.
            - Use head-to-head history if available to refine the analysis.
            - Avoid relying on external variables like weather or player-specific stats, as they are not provided.
            """
        else:
            return """
            ### General Specificity:
            - Use trends and recent performance data to refine predictions.
            - Focus on team-level patterns and overall scoring trends to improve consistency.
            """

    sport_specificity = switch(sport)

    GPT_prompt = f"""
    Ok chat, you are a highly skilled sports analyst specializing in the analysis of all sports. Your task is to evaluate and refine the given forecast based on an advanced analysis of the following data:

    ### Important Notes:
    - My initial forecast is based on **data sources and statistical models** that I cannot fully provide here (e.g., external trends, proprietary analysis).
    - While this forecast is **reliable and well-researched**, it is **not perfect** and does not reflect all situational variables (e.g., player injuries, live game conditions).

    ### Match Details:
    - **Home Team**: {match__['home_team']}
    - **Away Team**: {match__['away_team']}
    - **Date**: {match__['date']}
    - **Sport**: {match__['sport']}

    ### My Forecast:
    - **Prediction**: {match__['prediction']}
    - **Correct Score**: {match__['correct_score']}
    - **Average Score**: {match__['average_score']}

    ### Recent Team Performance:
    {last_match}
    - **Key Trends**:
    {trend}

    ### Task:
    1. Analyze the **consistency of the provided forecast** (`prediction` and `correct_score`) using probabilities, recent matches, and trends.
    2. Evaluate the **strength of my forecast** in light of the data provided and identify any gaps or inconsistencies.
    3. Provide a **realistic prediction** based on:
        - The team's recent performance (e.g., win/loss streaks, scoring trends).
        - Head-to-head performance.
        - Team-level trends (e.g., scoring patterns, defensive consistency).
    4. **If you are confident that one team will win and the recent matches and form confirm this, clearly state that the home or away team is the likely winner**.
    5. **If you have any doubts or see elements that might cause the prediction to change, mention them clearly**.
    6. Output the following in JSON format:
    ```json
    {{
        "home_probability_api": <float>,
        "draw_probability_api": <float>,
        "away_probability_api": <float>,
        "prediction_api": "<string>",
        "correct_score_api": "<string>",
        "average_score_api": <float>
    }}
    ```

    ### Additional Notes:
    {sport_specificity}
    """
    return GPT_prompt


def wait_for_element(driver, xpath, timeout=120):
    """
    Wait for an element to appear within the given timeout.
    Reload the page if the element is not found.
    """
    retries = 3  # Maximum retries
    for attempt in range(retries):
        try:
            return WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
        except Exception as e:
            if attempt < retries - 1:
                print(f"Element not found, reloading page... Attempt {attempt + 1}/{retries}")
                driver.refresh()
            else:
                print("Max retries reached. Element not found.")
                return None  # Fail gracefully

def analys_per_link(array, driver):
    matches = []
    analyse_manual = []
    filtered_array = [
        match__ for match__ in array
        if (
            (float(match__['initial_difference']) >= 38 and ("football" in str(match__['sport']).lower() or "football" in match__['link']))
            or (float(match__['initial_difference']) >= 45)
            or (float(match__['initial_difference']) >= 15 and "american" in str(match__['sport']).lower())
            or (float(match__['initial_difference']) >= 30  and ("rugby" in str(match__['sport']).lower() or "rugby" in match__['link']))
        )
    ]
    for match__ in filtered_array:
            link = match__['link'] if (match__ and len(match__['link'])>2) else None
            mots_interdits = ["ncaa", "chile", "tb2l", "u21", "georgi", "u19", "bulgar", "ligue-b", "espoir", "-a2-", "-a2/", "2-bundesliga"]
            paires_interdites = [("basket", "al-"), ("basket", "austr"), ("foot", "austr"), ("basket", "nbb"), ("basket", "mhl"), ("rugby", "women")]
            if link and all(mot not in link for mot in mots_interdits) and all(not (mot1 in link and mot2 in link) for mot1, mot2 in paires_interdites):
                try:
                    link = link.replace('https:/www.forebet.com/https:/www.forebet.com/', 'https:/www.forebet.com/')
                    print(f"Navigating to: {link}")
                    driver.get(link)
                    content_xpath = '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]'
                    content = wait_for_element(driver, content_xpath)
                    if not content:
                        print("Critical element missing. Skipping this page.")
                        return
                    div_xpaths = [
                        './/div[contains(@class, "st_scrblock")]',
                        './/div[contains(@class, "mx-width_hc")]'
                    ]
                    divs = []
                    for xpath in div_xpaths:
                        try:
                            divs = content.find_elements(By.XPATH, xpath)
                            if divs:
                                break
                        except Exception as e:
                            continue
                    
                    expected_url = driver.current_url
                    div_count = len(divs) if divs else 0
                    
                    print(f"Divs found: {div_count} at {expected_url}")
                    if divs:
                        div_count = len(divs)
                        first_divs = divs[:3] if div_count > 3 else divs
                        last_match = forebet_add_title_on_htmlElement(match__['home_team'], match__['away_team'], first_divs)
                        trend = clean_text(get_trend_forebet(driver))
                        if check_exists_by_xpath(driver, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]//div[contains(@class, "match_intro_tab")]') == 0:
                            trend += getinnertextXpath(driver, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]//div[contains(@class, "match_intro_tab")]')
                        #find result if present
                        if check_exists_by_xpath(driver, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]//div[@class="lscr_td"]//span') == 0:
                            final_score = getinnertextXpath(driver, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]//div[@class="lscr_td"]//span')
                        else:
                            final_score = ""
                            
                        GPT_prompt = prompt(match__, last_match, trend)
                        json_result = get_gpt_response_name(" ", GPT_prompt)
                        if json_result:
                            if len(json_result) > 0:
                                home_probability = to_percentage(match__['home_probability'])
                                draw_probability = to_percentage(match__['draw_probability'])
                                away_probability = to_percentage(match__['away_probability'])
                                average_score = match__['average_score']
                                home_probability_api = to_percentage(json_result["home_probability_api"]) if json_result["home_probability_api"] else ''
                                draw_probability_api = to_percentage(json_result["draw_probability_api"]) if json_result["draw_probability_api"] else ''
                                away_probability_api = to_percentage(json_result["away_probability_api"]) if json_result["away_probability_api"] else ''
                                average_score_api = json_result["average_score_api"] if json_result["average_score_api"] else ''
                                match_info = {
                                    'home_team': match__['home_team'],
                                    'away_team': match__['away_team'],
                                    'date': match__['date'],
                                    'home_probability': home_probability,
                                    'draw_probability': draw_probability,
                                    'away_probability': away_probability,
                                    'initial_difference': match__['initial_difference'],
                                    'initial_difference_api': abs(float(home_probability_api) - float(away_probability_api)),
                                    "home_probability_api": home_probability_api if home_probability_api else '',
                                    "draw_probability_api": draw_probability_api if draw_probability_api else '',
                                    "away_probability_api": away_probability_api if away_probability_api else '',
                                    'prediction': match__['prediction'],
                                    "prediction_api": json_result["prediction_api"] if json_result["prediction_api"] else '',
                                    'correct_score': match__['correct_score'],
                                    "correct_score_api": json_result["correct_score_api"] if json_result["correct_score_api"] else '',
                                    'average_score': average_score,
                                    "average_score_api": average_score_api if average_score_api else '',
                                    'sport': match__['sport'],
                                    "final_score": final_score,
                                    'link': match__['link']
                                }
                                append_new_line('analyse-log.txt', str(match_info))
                                matches.append(match_info)
                                check_and_refresh(driver, expected_url, timeout=120)
                        else:
                            append_new_line('content.txt', str(set_text(match__)))
                except Exception as e:
                    print(e)
                    append_new_line('error_by_link.txt', str(set_text(match__)))
                    append_new_line('error_by_link.txt', str(e))
                    analyse_manual.append(match__)
                    continue
    if len(matches) < 2 :
        if len(filtered_array) > 2:
            save_to_excel(filtered_array, "IA_forebet.xlsx")
    if len(analyse_manual) > 2 :
        save_to_excel(analyse_manual, "analyse-manual.xlsx")
    return matches

def forebet_scrap_history(driver):
    
    allcontent = str("")
    all_link = []
    driver.get("https://www.forebet.com/")
    waitloading(2, driver=driver)
    click_consent(driver, 'en')
    for ii in  ('https://www.forebet.com/en/football-tips-and-predictions-for-today', 
                'https://www.forebet.com/en/basketball/predictions-today',
                'https://www.forebet.com/en/hockey/predictions-today',
                'https://www.forebet.com/en/american-football/all-predictions',
                'https://www.forebet.com/en/volleyball/predictions-today',
                'https://www.forebet.com/en/handball/predictions-today',
                'https://www.forebet.com/en/rugby/predictions-today'):
        driver.get(ii)
        sport = str(ii.split("en/")[1].split("/")[0])
        sport = sport.split("-")[0] if '-' in sport else sport
        waitloading(2, driver=driver)
        if check_exists_by_xpath(driver, '//div[contains(@class, "fc-dialog-container")]//div[contains(@class, "fc-close fc-icon-button")]//span') == 0:
            tryAndRetryClickXpath(driver, '//div[contains(@class, "fc-dialog-container")]//div[contains(@class, "fc-close fc-icon-button")]//span')
        links = driver.find_elements(By.XPATH, '//div[contains(@class, "moduletable")]//div[contains(@class, "calpick")]//table[contains(@id, "cal-12")]//td[contains(@class, "active")]')
        if links and len(links) > 1:
            for el in links:
                link_ = findATTR(el, ".//a", "href")
                if link_:
                    all_link.append(link_)
                    append_new_line(r'forebet/all_link.html', str(link_))
    if all_link and len(all_link) > 1:
        for date in all_link:
            print(clean_text(date))
            driver.get(date)
            waitloading(1, driver=driver)
            content = driver.find_element(By.XPATH, '//table[contains(@class, "allcontent")]//td[contains(@class, "contentmiddle")]')
            content = (content.get_attribute('outerHTML'))
            if content:
                content = ajouter_sportfill(content, sport)
            allcontent += content
            append_new_line(r'forebet/history-log.html', str(content))
    if allcontent:
        return allcontent


def remplacer_en_mois_annee(texte):
    pattern = r'\b(\d{2})(\d{4})\b'
    texte_modifie = re.sub(pattern, r'\1/\2', texte)
    
    return texte_modifie

def forebet_add_title_on_htmlElement(home_team, away_team, first_divs):
    content_final, title = "", ""
    
    # Normalize the home_team and away_team by stripping extra spaces
    home_team_normalized = " ".join(home_team.lower().split())
    away_team_normalized = " ".join(away_team.lower().split())
    
    if first_divs:
        for i, block in enumerate(first_divs):
            text_extract = clean_html_and_return_innertext(block)
            
            # Normalize the block text
            text_extract_normalized = " ".join(text_extract.lower().split())

            # Count occurrences of normalized home_team and away_team
            count_word1 = text_extract_normalized.count(home_team_normalized)
            count_word2 = text_extract_normalized.count(away_team_normalized)
            
            # Determine the title based on the word counts
            if count_word1 > 0 and (count_word1 - 3) > count_word2:
                title = " . Recent Match History: " + str(home_team) + " Last Matches Overview : "
            elif count_word2 > 0 and (count_word2 - 3) > count_word1:
                title = " . Recent Match History: " + str(away_team) + " Last Matches Overview : "
            else:
                title = " . Comprehensive Statistics: " + str(home_team) + " and " + str(away_team) + " Analysis : "
            
            # Append the title and the block content
            content_final += title + text_extract

    return remplacer_en_mois_annee(content_final)
